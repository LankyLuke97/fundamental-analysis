import subprocess
import sys

subprocess.check_call([sys.executable, '-m', 'pip', 'install', '-U', '-r', 'requirements.txt'])

from collections import defaultdict
import csv
from datetime import datetime
from dateutil.relativedelta import relativedelta
from dotenv import load_dotenv, dotenv_values
import json
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import pandas as pd
from pathlib import Path
import requests
from sklearn import linear_model
from sklearn.linear_model import LinearRegression
from sklearn import metrics
from sklearn.metrics import mean_squared_error, mean_absolute_error
from sklearn.model_selection import train_test_split, cross_val_score
from time import perf_counter
from typing import NamedTuple

class Record(NamedTuple):
    ticker: str
    weight: float
    present_value: float
    margin_of_safety: float
    current_price: float

load_dotenv()
fmp_key = os.getenv("FMP_KEY")

def request_data(endpoint, **parameters):
    request = requests.get(f'{endpoint}?{"".join([f"{key}={value}&" for key, value in parameters.items()])}apikey={fmp_key}')
    if request.status_code == 200:
        return request.content
    print(f'Request to endpoint {endpoint} unsuccessful: {request.status_code}')
    return []

def request_data_to_json(json_file_path, endpoint, **parameters):
    param_mapping = {'from_': 'from'}
    for param, map_to in param_mapping.items():
        if param in parameters: parameters[map_to] = parameters.pop(param)
    if not up_to_date(json_file_path):
        json_data = json.loads(request_data(endpoint, **parameters))
        os.makedirs(Path(json_file_path).parent, exist_ok=True)
        with open(json_file_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=4)
    with open(json_file_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def up_to_date(filepath):
    time_stamp = datetime.fromtimestamp(os.path.getmtime(filepath)) if filepath.exists() else None
    return time_stamp and (time_stamp.year, time_stamp.month, time_stamp.day) == (datetime.today().year, datetime.today().month, datetime.today().day)

available_tickers_json = Path('data', 'available_tickers.json')
if not up_to_date(available_tickers_json):
    request_data_to_json(available_tickers_json, 'https://financialmodelingprep.com/api/v3/financial-statement-symbol-lists')

with open(available_tickers_json, 'r', encoding='utf-8') as f:
    available_tickers = json.load(f)

records = []
watchlist = sys.argv[1:]
if not watchlist:
    with open('watchlist.txt', 'r', encoding='utf-8') as f:
        watchlist = [ticker.strip() for ticker in f.readlines()]

linear_regression_df = defaultdict(list)

for ticker in watchlist if watchlist else sys.argv[1:]:
    if ticker not in available_tickers:
        print(f'{ticker} has no associated financial statements at financial modeling prep.')
        continue
    income_statement = request_data_to_json(Path('data', ticker, 'income_statements.json'),
                                            f'https://financialmodelingprep.com/api/v3/income-statement/{ticker}',
                                            period='annual',
                                            limit='11')
    balance_sheet = request_data_to_json(Path('data', ticker, 'balance_sheets.json'),
                                         f'https://financialmodelingprep.com/api/v3/balance-sheet-statement/{ticker}',
                                         period='annual',
                                         limit='11')
    _metrics = request_data_to_json(Path('data', ticker, 'metrics.json'),
                                  f'https://financialmodelingprep.com/api/v3/key-metrics/{ticker}',
                                  period='annual',
                                  limit='11')
    eod = request_data_to_json(Path('data', ticker, 'end_of_day.json'),
                                  f'https://financialmodelingprep.com/api/v3/historical-price-full/{ticker}',
                                  )
    dividends = request_data_to_json(Path('data', ticker, 'dividends.json'),
                                  f'https://financialmodelingprep.com/api/v3/historical-price-full/stock_dividend/{ticker}',
                                  )
    
    key_info_income = ['date',
                       'revenue',
                       'netIncome',
                       'epsdiluted',
                      ]
    key_info_balance = ['date',
                        'cashAndCashEquivalents',
                        'totalCurrentAssets',
                        'totalCurrentLiabilities',
                        'totalNonCurrentLiabilities',
                        'totalEquity',
                        ]
    key_info_metrics = ['date',
                        'debtToEquity',
                        'peRatio',
                        'pbRatio',
                        'roic',
                        'dividendYield',
                        ]
    key_info_dividends = ['date',
                          'adjDividend',
                          ] # Can't process this in the same way
    
    df = pd.merge(pd.DataFrame({key : [year[key] for year in income_statement] for key in key_info_income}),
                  pd.DataFrame({key : [year[key] for year in balance_sheet] for key in key_info_balance}),
                  on='date').set_index('date')

    for col in df.columns:
        current_value = df.iloc[0][col]
        growth_rate = np.power(current_value / df[col].iloc[1:], 1 / np.arange(1, len(df))) - 1
        growth_rate_prev = np.power(df.iloc[1][col] / df[col].iloc[2:], 1 / np.arange(2, len(df))) - 1
        df[f'{col}_cagr'] = np.concatenate(([0], growth_rate))
        df[f'{col}_cagr_prev'] = np.concatenate(([0,0], growth_rate_prev))
    
    last_filings = [income_statement[year]['fillingDate'] for year in [0,1]]
    filing_prices = [float(eod['historical'][i]['adjClose']) for i in range(730) if eod['historical'][i]['date'] in last_filings]
    year_return = (filing_prices[0] / filing_prices[1]) - 1

    for index, row in df.interpolate(method='linear', axis=0).reset_index().iterrows():
        if index < 2: continue
        for col, value in row.items():
            if '_cagr_prev' not in col: continue
            linear_regression_df[f'{col}_{index}'].append(value)
    linear_regression_df['year_return'].append(year_return)
    linear_regression_df['ticker'].append(ticker)
            
    df = df.merge(pd.DataFrame({key : [year[key] for year in _metrics] for key in key_info_metrics}).set_index('date'), left_index=True, right_index=True)
    df['roic_calc'] = df.netIncome / (df.totalNonCurrentLiabilities + df.totalEquity)
    df['roic_avg'] = df.roic.expanding().mean()

    REFERENCE_RATE = 0.1
    years = [(y, w) for y, w in [(1, 0.17),(3, 0.22),(5,0.27),(10,0.34)] if y < len(df)]
    adjust = sum([w for _, w in years])
    years = [(y, w / adjust) for y, w in years]
    check = [('totalEquity',0.27),('epsdiluted',0.22),('revenue',0.17)]
    df.reset_index(inplace=True)
    weight = (df.loc[[year for year, _ in years], [f'{c}_cagr' for c, _ in check]] - REFERENCE_RATE).mul(
        pd.Series({year: weight for year, weight in years}),
        axis="index"
    ).mul(
        pd.Series({f'{c}_cagr': weight for c, weight in check}),
        axis=1
    ).sum().sum() + (df.roic.mean() * 0.34)

    present_value = 0
    if df.totalEquity_cagr.iloc[len(df)-1] > 0 and df.epsdiluted.iloc[0] > 0:
        COMPOUND_YEARS = 10
        DISCOUNT_RATE = 0.15
        MAX_GROWTH_RATE = 0.2
        compound_factor = np.power(1 + min(df.totalEquity_cagr.iloc[len(df)-1], MAX_GROWTH_RATE), COMPOUND_YEARS)
        discount_factor = np.power(1 + DISCOUNT_RATE, -COMPOUND_YEARS)
        future_value = df.epsdiluted.iloc[0] * compound_factor * df.peRatio.iloc[0]
        present_value = future_value * discount_factor
    
    df.to_csv(Path('data', '_analysis', f'{ticker}.csv'))
    records.append(Record(ticker=ticker, weight=weight, present_value=present_value, margin_of_safety=(present_value/2), current_price=eod['historical'][0]['adjClose']))

    stats_to_display = [
        'roic',
        'revenue',
        'netIncome',
        'epsdiluted',
        'totalEquity',
        'cashAndCashEquivalents',
    ]
    cagrs_cols = ['roic_avg'] + [f'{c}_cagr' for c in stats_to_display if c != 'roic'] 
    cagr_df = df.loc[[v for v in df.index.values if v in [1, 3, 5, min(max(df.index.values), 10)]], cagrs_cols].transpose().iloc[:, ::-1]
    refined_df = df[stats_to_display].transpose().iloc[:, ::-1]
    ratios_df = df[key_info_metrics].drop(['date','roic'], axis=1)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in dataframe_to_rows(cagr_df, header=True, index=True):
        sheet.append(row)
    for row in sheet["A1:E11"]:
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value >= 0.1:
                cell.fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            elif isinstance(cell.value, (int, float)):
                cell.fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    sheet.move_range("A1:E11", rows=0, cols=12)
    for row in dataframe_to_rows(refined_df, header=True, index=True):
        sheet.append(row)
    for row in dataframe_to_rows(ratios_df.iloc[:1], index=False):
        sheet.append(row)
    sheet.move_range("A12:L21", rows=-11, cols=0)
    sheet.move_range("A3:P8", rows=-1, cols=0)
    workbook.save(Path('data', '_analysis', f'{ticker}.xlsx'))

'''
Tasks
Convert the writing of data to the excel directly to the correct location rather than just appending to the sheet:
- https://stackoverflow.com/questions/77914585/pandas-openpyxl-write-dataframe-with-left-corner-on-a-specific-cell

Add margin of safety calculations

Improve formatting to match existing (cleaner) style
'''
'''
records = sorted(records, key=lambda record: record.weight, reverse=True)
with open('watchlist_analysis.csv', mode="w", newline="", encoding="utf-8") as file:
    writer = csv.DictWriter(file, fieldnames=Record._fields)
    writer.writeheader()
    for record in records:
        writer.writerow(record._asdict())

linear_regression_df = pd.DataFrame(linear_regression_df).set_index('ticker')
x = linear_regression_df.drop(columns=['year_return'])
y = linear_regression_df['year_return']

X_train, X_test, y_train, y_test = train_test_split(x, y, test_size=0.25, random_state=100)
reg_model = linear_model.LinearRegression()
reg_model.fit(X_train, y_train)
y_pred = reg_model.predict(X_test)
reg_model_diff = pd.DataFrame({'Actual value': y_test, 'Predicted value': y_pred})

mae = mean_absolute_error(y_test, y_pred)
mse = mean_squared_error(y_test, y_pred)
r2 = np.sqrt(mean_squared_error(y_test, y_pred))

print(reg_model_diff)
print('Mean Absolute Error:', mae)
print('Mean Square Error:', mse)
print('Root Mean Square Error:', r2)
'''